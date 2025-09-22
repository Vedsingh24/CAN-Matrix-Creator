import tkinter as tk
from tkinter import filedialog, messagebox, PhotoImage
import cantools
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import sys
from ctypes import windll
from pathlib import Path
import ctypes

# DPI awareness (Windows)
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

def get_base_assets_path() -> Path:
    """
    Returns the base path for assets:
    - Uses `sys._MEIPASS` when running in a PyInstaller-built executable.
    - Falls back to the script's directory during development (e.g., PyCharm).
    """
    return Path(getattr(sys, "_MEIPASS", Path(__file__).parent))

def relative_to_assets(path: str) -> str:
    """Resolve a file path inside the assets folder."""
    base_path = get_base_assets_path()
    asset_path = base_path / path
    if not asset_path.exists():
        raise FileNotFoundError(f"Asset file not found: {asset_path}")
    return str(asset_path)

def process_dbc_file(file_path: str, output_file: str, silent: bool = False) -> str | None:
    """
    Convert a DBC file to a CAN Matrix Excel (.xlsx).
    Returns the output file path if success, None otherwise.
    """
    try:
        db = cantools.database.load_file(file_path)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CAN Messages"

        headers = [
            "CAN ID & Message Name", "Signal Name", "Byte Ordering", "Signed/Unsigned", "Start Bit", "Length",
            "Factor", "Offset", "Min Value", "Max Value", "Units"
        ]
        sheet.append(headers)

        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for col in range(1, len(headers) + 1):
            header_cell = sheet.cell(row=1, column=col)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cell.fill = header_fill

        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        current_row = 2

        for message in db.messages:
            can_id = hex(message.frame_id)
            message_name = message.name
            message_name_cell = f"{can_id} - {message_name}"

            for signal in message.signals:
                signal_name = signal.name
                start_bit = signal.start
                length = signal.length
                signed = "Signed" if signal.is_signed else "Unsigned"
                factor = signal.scale
                offset = signal.offset
                byte_ordering = "Motorola" if signal.byte_order == "big_endian" else "Intel"
                units = signal.unit if signal.unit else "N/A"
                min_value = signal.minimum if signal.minimum is not None else "N/A"
                max_value = signal.maximum if signal.maximum is not None else "N/A"

                row_data = [
                    message_name_cell, signal_name, byte_ordering, signed, start_bit,
                    length, factor, offset, min_value, max_value, units
                ]
                for c, data in enumerate(row_data, start=1):
                    cell = sheet.cell(row=current_row, column=c, value=data)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                current_row += 1

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in sheet.iter_rows(min_col=col, max_col=col, min_row=1, max_row=current_row - 1):
                val = row[0].value
                if val is not None:
                    max_length = max(max_length, len(str(val)))
            sheet.column_dimensions[column_letter].width = max_length + 2

        border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))
        for row in sheet.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border_style

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_file)

        if not silent:
            messagebox.showinfo("Success", f"Data written to {output_file}")
        return output_file
    except Exception as e:
        if not silent:
            messagebox.showerror("Error", f"An error occurred: {e}")
        else:
            print(f"[ERROR] {file_path} -> {e}")
        return None

def process_excel_to_dbc(excel_path: str, output_dbc: str, silent: bool = False) -> str | None:
    """
    Convert a CAN Matrix Excel (.xlsx) back to a DBC file.
    Returns the output DBC path if success, None otherwise.
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        sheet = wb.active

        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        expected = [
            "CAN ID & Message Name", "Signal Name", "Byte Ordering", "Signed/Unsigned", "Start Bit", "Length",
            "Factor", "Offset", "Min Value", "Max Value", "Units"
        ]
        if header[:len(expected)] != expected:
            message = "Excel format is not recognized. Expected CAN Matrix header."
            if not silent:
                messagebox.showerror("Error", message)
            else:
                print(f"[ERROR] {excel_path} -> {message}")
            return None

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        messages = {}  # (frame_id, msg_name) -> list[signal dict]

        for r in rows:
            if not r or all(v is None for v in r):
                continue

            can_msg, sig_name, byte_ordering, signed_text, start_bit, length, factor, offset, min_v, max_v, units = r[:11]
            if not can_msg or not sig_name:
                continue

            try:
                id_part, name_part = [p.strip() for p in str(can_msg).split("-", 1)]
                frame_id = int(id_part, 16)
                msg_name = name_part
            except Exception:
                message = f"Invalid 'CAN ID & Message Name' cell: {can_msg}"
                if not silent:
                    messagebox.showerror("Error", message)
                else:
                    print(f"[ERROR] {excel_path} -> {message}")
                return None

            byte_order = "big_endian" if str(byte_ordering).strip().lower().startswith("motorola") else "little_endian"
            is_signed = str(signed_text).strip().lower().startswith("signed")
            start = int(start_bit)
            leng = int(length)
            scale = float(factor) if factor is not None else 1.0
            offs = float(offset) if offset is not None else 0.0
            minimum = float(min_v) if (min_v not in (None, "N/A")) else None
            maximum = float(max_v) if (max_v not in (None, "N/A")) else None
            unit = None if (units in (None, "N/A")) else str(units)

            signals = messages.setdefault((frame_id, msg_name), [])
            signals.append({
                "name": str(sig_name),
                "start": start,
                "length": leng,
                "byte_order": byte_order,
                "is_signed": is_signed,
                "scale": scale,
                "offset": offs,
                "minimum": minimum,
                "maximum": maximum,
                "unit": unit,
            })

        db = cantools.database.Database()
        from cantools.database.can.signal import Signal
        from cantools.database.can.message import Message

        for (frame_id, msg_name), sigs in messages.items():
            ct_signals = [
                Signal(
                    name=s["name"],
                    start=s["start"],
                    length=s["length"],
                    byte_order=s["byte_order"],
                    is_signed=s["is_signed"],
                    scale=s["scale"],
                    offset=s["offset"],
                    minimum=s["minimum"],
                    maximum=s["maximum"],
                    unit=s["unit"],
                )
                for s in sigs
            ]

            if ct_signals:
                max_end_bit = max(s.start + s.length for s in ct_signals)
                dlc = max(1, min(8, (max_end_bit + 7) // 8))
            else:
                dlc = 8

            msg = Message(
                frame_id=frame_id,
                name=msg_name,
                length=dlc,
                signals=ct_signals,
                is_extended_frame=(frame_id > 0x7FF),
            )
            db.messages.append(msg)

        Path(output_dbc).parent.mkdir(parents=True, exist_ok=True)
        with open(output_dbc, "w", encoding="utf-8") as f:
            f.write(db.as_dbc_string())

        if not silent:
            messagebox.showinfo("Success", f"DBC created: {output_dbc}")
        else:
            print(f"[SAVED] {output_dbc}")
        return output_dbc
    except Exception as e:
        if not silent:
            messagebox.showerror("Error", f"Failed to create DBC: {e}")
        else:
            print(f"[ERROR] {excel_path} -> {e}")
        return None

def run_batch_dbc_to_excel(base_name: str):
    """
    Batch mode: select multiple DBC files; then choose an output directory.
    Saves each Excel as: [Base Name] CAN Matrix - [OriginalName].xlsx
    """
    file_paths = filedialog.askopenfilenames(
        title="Select DBC files",
        filetypes=[("DBC Files", "*.dbc"), ("All Files", "*.*")]
    )
    if not file_paths:
        return

    dst_dir = filedialog.askdirectory(title="Select output directory")
    if not dst_dir:
        return

    outdir = Path(dst_dir)
    try:
        outdir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        messagebox.showerror("Error", f"Cannot create output directory:\n{outdir}\n\n{e}")
        return

    base = (base_name or "").strip() or "Output"

    successes = 0
    failed = 0
    outputs = []

    for fp in file_paths:
        stem = Path(fp).stem
        out = outdir / f"{base} CAN Matrix - {stem}.xlsx"
        saved = process_dbc_file(str(fp), str(out), silent=True)
        if saved:
            successes += 1
            outputs.append(saved)
            print(f"[SAVED] {saved}")
        else:
            failed += 1

    msg = f"Batch complete.\nSaved: {successes}\nFailed: {failed}\nFolder: {outdir}"
    if outputs:
        msg += f"\n\nExample:\n{outputs[0]}"
    messagebox.showinfo("Batch Result", msg)

def run_single_dbc_to_excel():
    """
    Single-file DBC → Excel with Save As dialog.
    """
    file_path = filedialog.askopenfilename(
        title="Select a DBC file",
        filetypes=[("DBC Files", "*.dbc"), ("All Files", "*.*")]
    )
    if not file_path:
        return

    default_name = f"{Path(file_path).stem} CAN Matrix.xlsx"
    output_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not output_file:
        return

    process_dbc_file(file_path, output_file, silent=False)

def run_batch_excel_to_dbc(base_name: str):
    """
    Batch mode: select multiple CAN Matrix Excel files; then choose an output directory.
    Saves each DBC as: [Base Name] - [OriginalName].dbc
    """
    excel_paths = filedialog.askopenfilenames(
        title="Select CAN Matrix Excel files",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not excel_paths:
        return

    dst_dir = filedialog.askdirectory(title="Select output directory")
    if not dst_dir:
        return

    outdir = Path(dst_dir)
    try:
        outdir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        messagebox.showerror("Error", f"Cannot create output directory:\n{outdir}\n\n{e}")
        return

    base = (base_name or "").strip() or "Output"

    successes = 0
    failed = 0
    outputs = []

    for xp in excel_paths:
        stem = Path(xp).stem
        out = outdir / f"{base} - {stem}.dbc"
        saved = process_excel_to_dbc(str(xp), str(out), silent=True)
        if saved:
            successes += 1
            outputs.append(saved)
        else:
            failed += 1

    msg = f"Batch complete.\nSaved: {successes}\nFailed: {failed}\nFolder: {outdir}"
    if outputs:
        msg += f"\n\nExample:\n{outputs[0]}"
    messagebox.showinfo("Batch Result", msg)

def run_single_excel_to_dbc():
    """
    Single-file CAN Matrix (Excel) → DBC with Save As dialog.
    """
    excel_path = filedialog.askopenfilename(
        title="Select CAN Matrix (Excel)",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not excel_path:
        return

    default_name = f"{Path(excel_path).stem}.dbc"
    output_dbc = filedialog.asksaveasfilename(
        defaultextension=".dbc",
        initialfile=default_name,
        filetypes=[("DBC Files", "*.dbc"), ("All Files", "*.*")]
    )
    if not output_dbc:
        return

    process_excel_to_dbc(excel_path, output_dbc, silent=False)

# UI
root = tk.Tk()
root.title("DBC ↔ CAN Matrix Converter")
root.geometry("840x380")

# Set icon if available
try:
    icon_img = PhotoImage(file=relative_to_assets("KineticGreen.png"))
    root.wm_iconphoto(True, icon_img)
except Exception:
    pass

# Set AppUserModelID (Windows)
if sys.platform == "win32":
    try:
        app_id = "KineticGreen.UDS"
        windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
    except Exception:
        pass

title_label = tk.Label(root, text="Convert between DBC and CAN Matrix (Excel)", font=("Segoe UI", 11, "bold"))
title_label.pack(pady=10)

controls = tk.Frame(root)
controls.pack(pady=8)

batch_var = tk.BooleanVar(value=False)
batch_check = tk.Checkbutton(controls, text="Batch mode", variable=batch_var)
batch_check.grid(row=0, column=0, padx=8, pady=4, sticky="w")

tk.Label(controls, text="Base Name (used in batch):").grid(row=0, column=1, padx=8, pady=4, sticky="e")
base_name_var = tk.StringVar(value="")
base_name_entry = tk.Entry(controls, textvariable=base_name_var, width=34)
base_name_entry.grid(row=0, column=2, padx=8, pady=4, sticky="w")

btns = tk.Frame(root)
btns.pack(pady=16)

dbc_to_excel_btn = tk.Button(
    btns,
    text="DBC → CAN Matrix (Excel)",
    width=32,
    command=lambda: run_batch_dbc_to_excel(base_name_var.get()) if batch_var.get() else run_single_dbc_to_excel()
)
dbc_to_excel_btn.grid(row=0, column=0, padx=10, pady=8)

excel_to_dbc_btn = tk.Button(
    btns,
    text="CAN Matrix (Excel) → DBC",
    width=32,
    command=lambda: run_batch_excel_to_dbc(base_name_var.get()) if batch_var.get() else run_single_excel_to_dbc()
)
excel_to_dbc_btn.grid(row=0, column=1, padx=10, pady=8)

hint = tk.Label(
    root,
    text=(
        "Batch naming:\n"
        " - DBC → Excel: [Base Name] CAN Matrix - [OriginalName].xlsx\n"
        " - Excel → DBC: [Base Name] - [OriginalName].dbc"
    ),
    justify="center"
)
hint.pack(pady=6)

root.mainloop()